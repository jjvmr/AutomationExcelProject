import openpyxl

inv_file = openpyxl.load_workbook("coalpublic2013.xlsx")
mine_list = inv_file['Hist_Coal_Prod']

mines_per_supplier_region = {}
total_productivity_per_region = {}
production_under_10 = {}


# Listing each respective mine company with their respective Supplier Region
for mine_row in range(5, mine_list.max_row + 1):
    supplier_region = mine_list.cell(mine_row, 14).value
    production = mine_list.cell(mine_row, 15).value
    labour_hours = mine_list.cell(mine_row, 17).value
    mine_company = mine_list.cell(mine_row, 4).value
    productivity_info = mine_list.cell(mine_row, 18)

    # 1. calculation of number of mining companies per region
    if supplier_region in mines_per_supplier_region:
        current_num_mines = mines_per_supplier_region.get(supplier_region)     # getting value of a key
        mines_per_supplier_region[supplier_region] = current_num_mines + 1     # setting value of a key
    else:
        mines_per_supplier_region[supplier_region] = 1

    # 2. calculation total value of labour productivity per region (production per labour hours)
    # dividing total production value by labour hours
    if supplier_region in total_productivity_per_region:
        current_total_value = total_productivity_per_region.get(supplier_region)
        total_productivity_per_region[supplier_region] = int(current_total_value + (production / labour_hours))
    elif production == 0:
        total_productivity_per_region[supplier_region] = int(production + labour_hours)
    else:
        total_productivity_per_region[supplier_region] = int(production / labour_hours)

    # 3. logic with productivity less than 10
    if production < 1000:
        production_under_10[mine_company] = int(production)

    # value for new info per mining company
    if production != 0:
        productivity_info = production / labour_hours
    else:
        productivity_info = production + labour_hours

inv_file.save("coal_with_productivity.xlsx")
